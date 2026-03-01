from datetime import date
from pathlib import Path
from openpyxl import load_workbook

XLSX_PATH = Path("inputs/2026-02-24_noh_maternity_cashflow_model_24m.xlsx")
OUT_PATH = Path(f"outputs/{date.today()}_noh_maternity_cashflow_comments.txt")

def main():
    if not XLSX_PATH.exists():
        raise FileNotFoundError(f"Missing file: {XLSX_PATH.resolve()}")

    wb = load_workbook(XLSX_PATH, data_only=True)

    lines = []
    lines.append(f"Workbook: {XLSX_PATH.name}")
    lines.append("")

    # List sheets
    lines.append("SHEETS")
    for name in wb.sheetnames:
        lines.append(f"- {name}")
    lines.append("")

    # Extract cell comments
    lines.append("COMMENTS")
    comment_count = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.comment and cell.comment.text:
                    comment_count += 1
                    text = cell.comment.text.strip().replace("\r\n", "\n").replace("\r", "\n")
                    author = getattr(cell.comment, "author", "") or ""
                    value = "" if cell.value is None else str(cell.value)
                    lines.append(f"[{sheet_name}] {cell.coordinate}")
                    if author:
                        lines.append(f"Author: {author}")
                    lines.append(f"Cell value: {value}")
                    lines.append("Comment:")
                    lines.append(text)
                    lines.append("")

    if comment_count == 0:
        lines.append("No cell comments found via openpyxl.")
        lines.append("If your notes are in Excel 'Notes', 'Threaded comments', or plain text cells, tell me and we will extract those too.")
        lines.append("")

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text("\n".join(lines), encoding="utf-8")

    print(f"Done. Wrote: {OUT_PATH}")

if __name__ == "__main__":
    main()
