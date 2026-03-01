# Purpose: Extract monthly cashflow data from NOH_Maternity_Cashflow_Model_24m.xlsx
#          into a clean CSV for analysis.
# Input:   inputs/NOH_Maternity_Cashflow_Model_24m.xlsx (sheet: Model_24m, rows 28-52)
# Output:  outputs/YYYY-MM-DD_noh_maternity_cashflow_24m.csv
# Author:  NOH Expansion Workbench
# Date:    2026-03-01

import csv
from datetime import date
from pathlib import Path
from openpyxl import load_workbook

XLSX_PATH = Path("inputs/2026-02-24_noh_maternity_cashflow_model_24m.xlsx")
OUT_PATH = Path(f"outputs/{date.today()}_noh_maternity_cashflow_24m.csv")

COLUMNS = [
    "month", "births", "ma_share", "cash_share", "ma_cases", "cash_cases",
    "pr_cash", "pr_ma", "total_receipts", "noh_fee", "medicare_fee",
    "residual_pre_clinicians", "enrolments_active", "mo_sessions", "mo_cost",
    "cs_rate", "cs_cases", "anaes_cost", "ob_core_count", "ob_pool_total",
    "locum_reserve", "ob_core_share", "clinician_costs", "net_cashflow",
    "cumulative_cash"
]

def main():
    if not XLSX_PATH.exists():
        raise FileNotFoundError(f"Missing file: {XLSX_PATH.resolve()}")

    wb = load_workbook(XLSX_PATH, data_only=True)
    ws = wb["Model_24m"]

    # Data rows: 29-52 (row 28 is the header)
    DATA_START_ROW = 29
    DATA_END_ROW = 52
    COL_START = 1   # Column A
    COL_END = 25    # Column Y

    rows_written = 0
    rows_skipped = 0
    out_rows = []

    for row_idx in range(DATA_START_ROW, DATA_END_ROW + 1):
        row_values = [
            ws.cell(row=row_idx, column=col).value
            for col in range(COL_START, COL_END + 1)
        ]
        # Skip fully empty rows
        if all(v is None for v in row_values):
            rows_skipped += 1
            continue
        # Replace None with empty string
        row_values = ["" if v is None else round(v, 4) if isinstance(v, float) else v for v in row_values]
        out_rows.append(row_values)
        rows_written += 1

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_PATH, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(COLUMNS)
        writer.writerows(out_rows)

    # Validation summary
    print("=" * 50)
    print("VALIDATION SUMMARY")
    print("=" * 50)
    print(f"Input file:    {XLSX_PATH}")
    print(f"Output file:   {OUT_PATH}")
    print(f"Rows written:  {rows_written}")
    print(f"Rows skipped:  {rows_skipped} (empty)")
    print(f"Columns:       {len(COLUMNS)}")
    print(f"Status:        {'PASS' if rows_written == 24 else 'WARN — expected 24 rows'}")
    print("=" * 50)

if __name__ == "__main__":
    main()
